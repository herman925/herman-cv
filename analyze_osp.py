import zipfile
import os
import sys

file_path = r'c:\Users\Herman\Downloads\my-cv\Backends\Master Transfer Form.xlsm'

print(f"Analyzing: {file_path}")

if not os.path.exists(file_path):
    print("File not found!")
    sys.exit(1)

# 1. Check for VBA Project via Zip
try:
    with zipfile.ZipFile(file_path, 'r') as z:
        files = z.namelist()
        if 'xl/vbaProject.bin' in files:
            print("\n[VBA Structure]")
            print(" - vbaProject.bin found (Contains Macros/VBA Code)")
        
        print("\n[Internal Structure (Top Level)]")
        for f in files:
            if f.count('/') <= 1:
                print(f" - {f}")
except Exception as e:
    print(f"Error reading zip structure: {e}")

# 2. Try to read Sheet Names and Headers using openpyxl
try:
    import openpyxl
    print("\n[Workbook Data Model]")
    wb = openpyxl.load_workbook(file_path, read_only=True, keep_vba=True)
    
    print(f"Sheets: {wb.sheetnames}")
    
    print("\n[Sheet Headers (Schema)]")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Get first row
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = list(row)
            break
        
        # Filter None values
        headers = [h for h in headers if h]
        if headers:
            print(f" - {sheet_name}: {headers[:10]} {'...' if len(headers) > 10 else ''}")
        else:
            print(f" - {sheet_name}: (Empty or No Headers)")

    print("\n[Defined Names (Named Ranges)]")
    if wb.defined_names:
        for dn in list(wb.defined_names.definedName)[:20]: # Limit to 20
            print(f" - {dn.name}")
            
except ImportError:
    print("\nopenpyxl not installed. Cannot read sheet content deeply.")
except Exception as e:
    print(f"\nError reading workbook with openpyxl: {e}")
